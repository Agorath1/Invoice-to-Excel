# The program converts Storefront PDFs to a excel spreadsheet
#
#Version      Author
#  1.1      Robertp3001

import fitz
import pandas
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import re
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import os
import json

print("Storefront Invoice to Excel.")


# Imports json data
def import_json(json_file):
    with open(json_file, 'r') as file:
        json_data = json.load(file)
    return json_data


# Opens the dialog box to select pdfs
def select_multiple_pdfs():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    print("Selecting files to convert.")
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    return file_paths


def auto_adjust_column(workbook):
    # Adjust column sizes of of every used column in the workbook

    # Style used for money columns
    currency_style = NamedStyle(name="currency_style", number_format="$#,##0.00")

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name

            for cell in col:
                if cell.row > 1:
                    # Checks if for columns that need to be converted to money format
                    if sheet_name == settings_info["Sheet1"] and column in settings_info["columns1_currency"]:
                        cell.style = currency_style
                    if sheet_name == settings_info["Sheet2"] and column in settings_info["columns2_currency"]:
                        cell.style = currency_style

                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            adjusted_width = (max_length + 1) * 1.2
            if col[0].column == columns1.index("AWGSELL") + 1:
                adjusted_width = 11.29
            ws.column_dimensions[column].width = adjusted_width


def workbook_from_path(excel_path):
    # Load the workbook and adjust column widths for all sheets
    wb = load_workbook(excel_path)
    auto_adjust_column(wb)

    # Save the workbook
    wb.save(excel_path)

    # Use the start command to open the file with its default program
    # print("Opening Excel document " + excel_path.split("/")[-1])
    # os.system(f'start "" "{excel_path}"')


# Load Global Settings
settings_info = import_json('settings.json')

# Store Global departments
departments = settings_info["departments"]
columns1 = settings_info["columns1"]
columns2 = list(columns1[0:5]) + settings_info["columns2"]

# Select Multiple Files
pdf_file_paths = select_multiple_pdfs()

for pdf_path in pdf_file_paths:
    pdf_document = fitz.open(pdf_path)

    data = []
    data2 = []

    sheet1_line = []
    for i in range(len(columns1)):
        sheet1_line.append("")

    sheet2_line = []
    counted_data = [0, 0, 0]

    # Extract text from each page and process
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        text = page.get_text("text")
        text = re.sub(r'\s+', ' ', text).split()

        counter = -1
        counter2 = -1
        while counter < len(text) - 1:
            counter += 1

            if text[counter] == "STORE":
                try:
                    sheet1_line[columns1.index("STORE")] = int(text[counter + 1])
                    counter += 2
                except:
                    pass
            if text[counter] == "DEPT" and len(text[counter + 1]) == 3:
                try:
                    sheet1_line[columns1.index("DEPT")] = departments[text[counter + 1][0:2]]
                except KeyError:
                    sheet1_line[columns1.index("DEPT")] = int(text[counter + 1][0:2])
                counter += 2
            if text[counter] == "DT:":
                sheet1_line[columns1.index("DELV DT")] = text[counter + 1]
                counter += 2
            if text[counter] == "INVOICE#":
                if ((sheet1_line[columns1.index("INVOICE#")] != int(text[counter + 1])) and
                     sheet1_line[columns1.index("INVOICE#")] != ""):
                    if not sheet2_line:
                        sheet2_line.append(0)
                        sheet2_line.append(0)
                        sheet2_line.append(0)
                    sheet2_line.append(counted_data[0])
                    sheet2_line.append(counted_data[1])
                    sheet2_line.append(counted_data[2])
                    data2.append(
                        list(data[-1][0:3]) +
                        [sheet1_line[columns1.index("PAGE")]] +
                        [data[-1][4]] + sheet2_line
                    )
                    print(
                        "Invoice " +
                        str(sheet1_line[columns1.index("DEPT")]) +
                        " " +
                        str(sheet1_line[columns1.index("INVOICE#")]) +
                        " completed."
                    )
                    sheet2_line = []
                    counted_data = [0, 0, 0]

                sheet1_line[columns1.index("INVOICE#")] = int(text[counter + 1])
                counter += 2
            if text[counter] == "PAGE" and (not text[counter + 1].isalpha()):
                sheet1_line[columns1.index("PAGE")] = int(text[counter + 1])
                counter += 2
            if text[counter] == "TOTAL" and text[counter + 1] == "ORDERED":
                sheet2_line.append(int(text[counter + 3]))
                counter += 4
            if text[counter] == "TOTAL" and text[counter + 1] == "SHIPPED":
                sheet2_line.append(int(text[counter + 3]))
                counter += 4
            if text[counter] == "INVOICE" and text[counter + 1] == "AMOUNT":
                sheet2_line.append(float(text[counter + 2].replace(',', '')))
                counter += 3

            try:
                if (text[counter].isdigit() and
                        text[counter + 1].isdigit() and
                        ("-" in text[counter + 2] or text[counter + 2] == "ITEM")):

                    for i in range(5, len(sheet1_line)):
                        sheet1_line[i] = ""

                    if text[counter - 1] == "PB":
                        sheet1_line[columns1.index("PB")] = text[counter - 1]

                    # Quantity Order
                    sheet1_line[columns1.index("QTY ORD")] = int(text[counter])
                    counted_data[0] += int(text[counter])
                    counter += 1
                    # Quantity Shipped
                    sheet1_line[columns1.index("QTY SHP")] = int(text[counter])
                    counted_data[1] += int(text[counter])
                    counter += 1
                    # Item Code
                    i = 0
                    if "-" in text[counter]:
                        sheet1_line[columns1.index("ITEM #")] = int(text[counter].split("-")[1])
                        counter += 1
                    else:
                        while True:
                            if not text[counter + i].isdigit():
                                i += 1
                            else:
                                break
                        sheet1_line[columns1.index("DESCRIPTION")] = data[-1][columns1.index("DESCRIPTION")]
                        sheet1_line[columns1.index("PRODUCT UPC")] = data[-1][columns1.index("PRODUCT UPC")]
                        sheet1_line[columns1.index("AWGSELL")] = " ".join(text[counter:(counter + i)]).replace(":", "")
                        sheet1_line[columns1.index("ITEM #")] = int(text[counter + i])
                        counter += i
                        data.append(list(sheet1_line))
                        continue

                    if sheet1_line[columns1.index("ITEM #")] == 320218:
                        counter = counter

                    i = 0
                    try:
                        while text[counter + i][0] != "0" or len(text[counter + i]) != 15:
                            i += 1
                    except IndexError:
                        pass
                    sheet1_line[columns1.index("DESCRIPTION")] = " ".join(text[counter:(counter + i)])
                    counter += i

                    sheet1_line[columns1.index("PRODUCT UPC")] = int(text[counter])
                    counter += 1

                    # Checks for not shipped.
                    i = 0
                    while True:
                        try:
                            float(text[counter + i])
                            break
                        except ValueError:
                            i += 1
                        except IndexError:
                            break
                    if i != 0:

                        sheet1_line[columns1.index("AWGSELL")] = " ".join(text[counter:counter + i])
                        counter += i

                    else:

                        sheet1_line[columns1.index("AWGSELL")] = float(text[counter])
                        counter += 1

                        sheet1_line[columns1.index("TOTAL ALLOW")] = float(text[counter])
                        counter += 1

                        sheet1_line[columns1.index("NET COST")] = float(text[counter])
                        counter += 2

                        try:
                            if str(int(text[counter])) == text[counter]:
                                sheet1_line[columns1.index("PACK")] = int(text[counter])
                                counter += 1
                        except ValueError:
                            pass

                        try:
                            float(text[counter + 1])
                            sheet1_line[columns1.index("UNT COST")] = float(text[counter])
                            counter += 1
                        except ValueError:
                            pass

                        sheet1_line[columns1.index("EXT NT COST")] = float(text[counter])
                        counted_data[2] += float(text[counter])
                        counter += 1

                        if text[counter] == "PB":
                            sheet1_line[columns1.index("PB")] = text[counter]
                            counter += 1

                        while True:
                            try:
                                float(text[counter])
                                sheet1_line[columns1.index("FREIGHT")] = float(text[counter])
                                break
                            except ValueError:
                                counter += 1
                        counter += 1

                        if "ITEM" in text[counter:counter + 5] and "OUT" in text[counter:counter + 6]:
                            counter2 = counter

                        if "WEIGHT:" in text[counter:counter + 20]:
                            while text[counter] != "WEIGHT:":
                                counter += 1
                            sheet1_line[columns1.index("TOTAL WEIGHT")] = float(text[counter + 1].replace(",", ""))
                            counter += 2

                    counter -= 1
                    data.append(list(sheet1_line))
                    if counter2 != -1:
                        counter = counter2
                        counter2 = -1
            except IndexError:
                pass

    if not sheet2_line:
        sheet2_line.append(0)
        sheet2_line.append(0)
        sheet2_line.append(0)
    sheet2_line.append(counted_data[0])
    sheet2_line.append(counted_data[1])
    sheet2_line.append(counted_data[2])
    data2.append(list(sheet1_line[0:5]) + sheet2_line)
    print("Invoice " + str(sheet1_line[columns1.index("DEPT")]) + " " + str(
        sheet1_line[columns1.index("INVOICE#")]) + " completed.")

    # Convert list to DataFrame
    df1 = pandas.DataFrame(data, columns=columns1)
    df2 = pandas.DataFrame(data2, columns=columns2)

    # Save DataFrame to Excel file
    excel_path = settings_info["new_path"] + pdf_path.split("/")[-1].split(".")[0]+ ".xlsx"
    print(excel_path)

    with pandas.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name=settings_info["Sheet1"], index=False)
        df2.to_excel(writer, sheet_name=settings_info["Sheet2"], index=False)

    workbook_from_path(excel_path)

print("Program complete.")
